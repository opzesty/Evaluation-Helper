from app import app
from flask import render_template, request, session, redirect, send_from_directory, make_response
from flask_session import Session
import os
import io

app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
app.config["UPLOAD_FOLDER"] = os.path.join(app.root_path, "uploads")
Session(app)

test_ip = "3.21.193.76" #test matt
production_ip = "3.14.124.94" #production matt

active_ip = test_ip

#@app.route('/test', methods = ['POST'])
#def test():
#    return make_response(request.form.get("name"))

@app.route('/')
def home():
    if not session.get("user"):
        return redirect("/login")
    if active_ip == production_ip:
        server_type = "production"
    else:
        server_type = "test"
    return render_template("index.html", server = server_type)

@app.route('/login', methods = ['POST', 'GET'])
def login():
    if request.method =='POST':
        session["user"] = request.form.get("user")
        session["password"] = request.form.get("password")
        session["team"] = "TEAM" + request.form.get("team_number")
        return redirect("/")
    return render_template('login.html')

@app.route('/logout', methods = ['GET'])
def logout():
    session.clear()
    return redirect("/")

@app.route('/pull_day_yaml', methods = ['POST'])
def pull_daily_observations_yaml():
    import requests
    import json
    import yaml

    session["day"] = request.form.get("day")

    r_session = requests.Session()
    r_session.post('https://' + active_ip + '/login', data={"username": session.get("user"), "password": session.get("password")}, verify=False)
    msel_r = r_session.get('https://' + active_ip + '/api/measure-evaluations', verify=False)
    msel_json = json.loads(msel_r.text)

    relevant_grading_opportunity = []
    for entry in msel_json:
        if entry["startDay"] == int(session["day"]) and entry["team"] == session["team"]:
            relevant_grading_opportunity.append(entry)

    relevant_grading_opportunity = sorted(relevant_grading_opportunity, key=lambda x: x['mselId'])

    matt_responses = []
    matt_responses.append('---')
    for entry in relevant_grading_opportunity:
        matt_responses.append('- inject_id: ' + entry['mselId'])
        matt_responses.append('  Inject Title: ' + entry['title'])
        matt_responses.append('  measure_code: ' + entry['measureCode'])
        matt_responses.append('  MOP/MOE Description: ' + entry['description'])
        matt_responses.append('  grade: ')
        matt_responses.append('  comment: >')
        matt_responses.append('\n')

    resp = make_response(render_template('observation_results.yaml', matt_responses = matt_responses))
    resp.headers['Content-Type'] = 'application/octet-stream'
    return resp

@app.route('/pull_day_excel', methods = ['POST'])
def pull_daily_observations_excel():
    import requests
    import json
    import yaml
    import xlsxwriter

    session["day"] = request.form.get("day")

    r_session = requests.Session()
    r_session.post('https://' + active_ip + '/login', data={"username": session.get("user"), "password": session.get("password")}, verify=False)

    msel_r = r_session.get('https://' + active_ip + '/api/measure-evaluations', verify=False)
    msel_json = json.loads(msel_r.text)

    relevant_grading_opportunity = []
    for entry in msel_json:
        if entry["startDay"] == int(session["day"]) and entry["team"] == session["team"]:
            relevant_grading_opportunity.append(entry)

    relevant_grading_opportunity = sorted(relevant_grading_opportunity, key=lambda x: x['mselId'])

    file_name = session["day"] + ".xlsx"
    workbook = xlsxwriter.Workbook(os.path.join(app.config["UPLOAD_FOLDER"], file_name))
    worksheet = workbook.add_worksheet()

    matt_responses = []
    current_msel = ''
    row_number = 1
    for entry in relevant_grading_opportunity:
        if current_msel == entry['mselId']:
            worksheet.write('B'+str(row_number), entry["measureCode"])
            worksheet.write('C'+str(row_number), entry["description"])
            worksheet.write('D'+str(row_number), entry["status"])
            worksheet.write('E'+str(row_number), entry["tacticalAssessmentComments"])
            row_number += 1
        else:
            current_msel = entry['mselId']
            worksheet.write('A'+str(row_number), "Inject ID:")
            worksheet.write('B'+str(row_number), entry["mselId"])
            worksheet.write('C'+str(row_number), "Inject Title:")
            worksheet.write('D'+str(row_number), entry["title"])
            row_number += 1
            worksheet.write('B'+str(row_number), "Measure Code")
            worksheet.write('C'+str(row_number), "MOP/MOE Description")
            worksheet.write('D'+str(row_number), "Grade")
            worksheet.write('E'+str(row_number), "Comment")
            row_number += 1
            worksheet.write('B'+str(row_number), entry["measureCode"])
            worksheet.write('C'+str(row_number), entry["description"])
            worksheet.write('D'+str(row_number), entry["status"])
            worksheet.write('E'+str(row_number), entry["tacticalAssessmentComments"])
            row_number += 1
 
    workbook.close()

    #resp = make_response(./session["day"] + ".xlsx")
    #resp.headers['Content-Type'] = 'text/xlsx'
    return send_from_directory(app.config["UPLOAD_FOLDER"], file_name, as_attachment=True)

@app.route('/pull_all_excel', methods = ['GET'])
def pull_all_observations_excel():
    import requests
    import json
    import yaml
    import xlsxwriter

    r_session = requests.Session()
    r_session.post('https://' + active_ip + '/login', data={"username": session.get("user"), "password": session.get("password")}, verify=False)

    msel_r = r_session.get('https://' + active_ip + '/api/measure-evaluations', verify=False)
    msel_json = json.loads(msel_r.text)

    relevant_grading_opportunity = []
    for entry in msel_json:
        if entry["team"] == session["team"]:
            relevant_grading_opportunity.append(entry)

    relevant_grading_opportunity = sorted(relevant_grading_opportunity, key=lambda x: x['mselId'])

    file_name = session["team"] + ".xlsx"
    workbook = xlsxwriter.Workbook(os.path.join(app.config["UPLOAD_FOLDER"], file_name))
    worksheet = workbook.add_worksheet()

    matt_responses = []
    current_msel = ''
    row_number = 1
    for entry in relevant_grading_opportunity:
        if current_msel == entry['mselId']:
            worksheet.write('B'+str(row_number), entry["measureCode"])
            worksheet.write('C'+str(row_number), entry["description"])
            worksheet.write('D'+str(row_number), entry["status"])
            worksheet.write('E'+str(row_number), entry["tacticalAssessmentComments"])
            row_number += 1
        else:
            current_msel = entry['mselId']
            worksheet.write('A'+str(row_number), "Inject ID:")
            worksheet.write('B'+str(row_number), entry["mselId"])
            worksheet.write('C'+str(row_number), "Inject Title:")
            worksheet.write('D'+str(row_number), entry["title"])
            row_number += 1
            worksheet.write('B'+str(row_number), "Measure Code")
            worksheet.write('C'+str(row_number), "MOP/MOE Description")
            worksheet.write('D'+str(row_number), "Grade")
            worksheet.write('E'+str(row_number), "Comment")
            row_number += 1
            worksheet.write('B'+str(row_number), entry["measureCode"])
            worksheet.write('C'+str(row_number), entry["description"])
            worksheet.write('D'+str(row_number), entry["status"])
            worksheet.write('E'+str(row_number), entry["tacticalAssessmentComments"])
            row_number += 1
 
    workbook.close()

    #resp = make_response(./session["day"] + ".xlsx")
    #resp.headers['Content-Type'] = 'text/xlsx'
    return send_from_directory(app.config["UPLOAD_FOLDER"], file_name, as_attachment=True)

@app.route('/update_msel', methods = ['POST'])
def send_observations():
    import requests
    import json
    from openpyxl import load_workbook

    evaluations = request.files['file']
    saved_file = evaluations.save(os.path.join(app.config["UPLOAD_FOLDER"], evaluations.filename))

    wb = load_workbook(filename = os.path.join(app.config["UPLOAD_FOLDER"], evaluations.filename))
    sheet = wb.active
    
    uploads = []
    inject_id = ""
    for row in sheet:
        if row[0].value != None and row[0].value.strip() == 'Inject ID:':
            inject_id = row[1].value
        elif row[1].value.strip() == 'Measure Code':
            pass
        else:
            entry = {"inject_id": inject_id, "measure_code": row[1].value, "grade": row[3].value, "comment": row[4].value}
            uploads.append(entry)

    r_session = requests.Session()

    r_session.post('https://' + active_ip + '/login', data={"username": session.get("user"), "password": session.get("password")}, verify=False)

    msel_r = r_session.get('https://' + active_ip + '/api/measure-evaluations', verify=False)

    msel_json = json.loads(msel_r.text)

    matt_responses = []

    for entry in msel_json:
        if not uploads:
            matt_responses.append("No changes to upload")
            break
        for evaluation in uploads:
            if entry['mselId'] == evaluation["inject_id"] and entry["measureCode"] == evaluation["measure_code"] and entry["team"] == session["team"]:
                if entry["status"] == evaluation["grade"] and entry["tacticalAssessmentComments"] == evaluation["comment"]:
                    pass
                else:
                    response = r_session.post(
                        'https://' + active_ip + '/api/measure-evaluations/update',
                        json={"id": entry['id'],
                        "status": evaluation["grade"],
                        "tacticalAssessmentComments": evaluation["comment"],
                        "operationalAssessmentComments": None},
                        headers={'Content-type': 'application/json; charset=utf-8'},
                        verify=False)
                    matt_responses.append(
                        "Observation {} for MSEL {}, measurecode {}, team {} was successfully changed to {}"
                        .format(entry['id'], entry['mselId'], entry['measureCode'], entry['team'], evaluation['grade']))
                    break

    return render_template('observation_results.html', matt_responses = matt_responses)
